import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

### READING THE SETS EXCEL FILE
file_sets = r"1_SETS.xlsx"
df_sets = pd.read_excel(file_sets, 'Sets', header=0,
                        keep_default_na='True', dtype='str')

### READING THE FIRST PAGE OF THE PARAMETERS EXCEL FILE - PARAMETERS & WRITING IT
file_parameters = r'2_PARAMETERS.xlsx'
df_par = pd.read_excel(file_parameters, 'Index')

### WRITING THE FILE
wb = openpyxl.load_workbook(file_parameters) # loading the excel file
dfp = None # defining an empty dataframe

for index, row in df_par.iterrows(): # for loop over the list of input parameters
    if row['Short name'] in wb.sheetnames:  # if the relative worksheet already exists, delete it
        del wb[row['Short name']]
    ws = wb.create_sheet(row['Short name'])  # create the relative worksheet
    n_sets = sum(pd.notnull(df_par.iloc[index, 5:]).tolist()) # calculate the number of SETS for that parameter
    if dfp is not None:
        del dfp
    if n_sets == 5: # if the number of SETS is equal to 5
        dfp = pd.DataFrame(
            columns=[df_par.iloc[index, 5], df_par.iloc[index, 6], df_par.iloc[index, 7], df_par.iloc[index, 8],
                     df_par.iloc[index, 9], 'Value'])
        for aa in df_sets[df_par.iloc[index, 5]].dropna(): # 5 annidated for loops over the different SETS
            for bb in df_sets[df_par.iloc[index, 6]].dropna():
                for cc in df_sets[df_par.iloc[index, 7]].dropna():
                    for dd in df_sets[df_par.iloc[index, 8]].dropna():
                        for ee in df_sets[df_par.iloc[index, 9]].dropna():
                            dfp = dfp.append(
                                {df_par.iloc[index, 5]: aa, df_par.iloc[index, 6]: bb, df_par.iloc[index, 7]: cc,
                                 df_par.iloc[index, 8]: dd, df_par.iloc[index, 9]: ee, 'Value': row['Default value']},
                                ignore_index=True) # inserting the elements of the input table of the parameter
    elif n_sets == 4: # if the number of SETS is equal to 4
        dfp = pd.DataFrame(
            columns=[df_par.iloc[index, 5], df_par.iloc[index, 6], df_par.iloc[index, 7], df_par.iloc[index, 8],
                     'Value'])
        for aa in df_sets[df_par.iloc[index, 5]].dropna(): # 4 annidated for loops over the different SETS
            for bb in df_sets[df_par.iloc[index, 6]].dropna():
                for cc in df_sets[df_par.iloc[index, 7]].dropna():
                    for dd in df_sets[df_par.iloc[index, 8]].dropna():
                        dfp = dfp.append(
                            {df_par.iloc[index, 5]: aa, df_par.iloc[index, 6]: bb, df_par.iloc[index, 7]: cc,
                             df_par.iloc[index, 8]: dd, 'Value': row['Default value']},
                            ignore_index=True) # inserting the elements of the input table of the parameter
    elif n_sets == 3: # if the number of SETS is equal to 3
        dfp = pd.DataFrame(
            columns=[df_par.iloc[index, 5], df_par.iloc[index, 6], df_par.iloc[index, 7], 'Value'])
        for aa in df_sets[df_par.iloc[index, 5]].dropna(): # 3 annidated for loops over the different SETS
            for bb in df_sets[df_par.iloc[index, 6]].dropna():
                for cc in df_sets[df_par.iloc[index, 7]].dropna():
                    dfp = dfp.append(
                        {df_par.iloc[index, 5]: aa, df_par.iloc[index, 6]: bb, df_par.iloc[index, 7]: cc,
                         'Value': row['Default value']},
                        ignore_index=True) # inserting the elements of the input table of the parameter
    elif n_sets == 2: # if the number of SETS is equal to 2
        dfp = pd.DataFrame(
            columns=[df_par.iloc[index, 5], df_par.iloc[index, 6], 'Value'])
        for aa in df_sets[df_par.iloc[index, 5]].dropna(): # 2 annidated for loops over the different SETS
            for bb in df_sets[df_par.iloc[index, 6]].dropna():
                dfp = dfp.append(
                    {df_par.iloc[index, 5]: aa, df_par.iloc[index, 6]: bb, 'Value': row['Default value']},
                    ignore_index=True) # inserting the elements of the input table of the parameter
    elif n_sets == 1: # if the number of SETS is equal to 1
        dfp = pd.DataFrame(
            columns=[df_par.iloc[index, 5], 'Value'])
        for aa in df_sets[df_par.iloc[index, 5]].dropna(): # 1 for loops over the different SETS
            dfp = dfp.append(
                {df_par.iloc[index, 5]: aa, 'Value': row['Default value']},
                ignore_index=True) # inserting the elements of the input table of the parameter

    for r in dataframe_to_rows(dfp, index=False, header=True): # writing all the rows in the excel sheet
        ws.append(r)
    for cell in ws['1']: # formatting in bold the first line
        cell.style = 'Pandas'
    ws.freeze_panes = ws['A2'] # freezing the panes in order to make scroll nicer
    ws.column_dimensions['A'].width = 27 # defining the column width
    ws.column_dimensions['B'].width = 27 # defining the column width
    ws.column_dimensions['C'].width = 27 # defining the column width
    ws.column_dimensions['D'].width = 27 # defining the column width
    ws.column_dimensions['E'].width = 27 # defining the column width

wb.save(file_parameters) # saving the excel file
